import openpyxl




def getRowCount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColCount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column

def getCelldata(path,sheetname,rowno,colno):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowno,column=colno).value

def setCellData(path,sheetname,rowno,colno,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rowno, column=colno).value = data
    workbook.save("/Volumes/Macintosh HD/For Mac/python project/pythongrid/ReadingExcel/testdata.xlsx")



path="/Volumes/Macintosh HD/For Mac/python project/pythongrid/ReadingExcel/testdata.xlsx"

sheetname = "Sheet1"

rows = getRowCount(path,sheetname)
cols = getColCount(path,sheetname)

print(rows,"---",cols)

print(getCelldata(path,sheetname,2,1))

print(setCellData(path,sheetname,1,4,"DOB"))


