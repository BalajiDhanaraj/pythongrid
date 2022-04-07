import openpyxl


workbook = openpyxl.load_workbook("/Volumes/Macintosh HD/For Mac/python project/pythongrid/ReadingExcel/testdata.xlsx")

sheet = workbook["Sheet1"]

totalrows  = sheet.max_row
totalcols  = sheet.max_column

print("total rows are : ",str(totalrows),"and total cols are : ",str(totalcols))


for rows in range(1,totalrows+1):
    for cols in range(1,totalcols+1):
        print(sheet.cell(row=rows,column=cols).value,end=" ")
    print()










