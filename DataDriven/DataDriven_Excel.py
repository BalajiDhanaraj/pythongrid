import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from selenium import webdriver


# parallel test

@pytest.fixture()
def log_on_failure(request,get_browser):
    yield
    item = request.node
    driver = get_browser
    if item.rep_call.failed:
        allure.attach(driver.get_screenshot_as_png(), name="dologin", attachment_type=AttachmentType.PNG)


def get_data():
    # return [
    #     ("trainer@way2automation.com", "kjsdfbksdf"),
    #     ("java@way2automation.com", "sdf"),
    #     ("info@way2automation.com", "sdfsdf")
    # ]
    workbook = openpyxl.load_workbook("/Volumes/Macintosh HD/For Mac/python project/pythongrid/DataDriven/testdata.xlsx")
    sheet = workbook['Sheet1']
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
            data = sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainList.insert(i,dataList)
        print(mainList)
    return mainList



def setup_function():
    global driver
    driver = webdriver.Chrome(
        executable_path="/Volumes/Macintosh HD/For Mac/python project/Browserdrivers/chromedriver")
    driver.get("http://facebook.com")


def teardown_function():
    driver.quit()

@pytest.mark.usefixtures("log_on_failure")
@pytest.mark.parametrize("username,password", get_data())
def test_dologin(username, password, get_browser):
    driver = get_browser
    driver.find_element_by_id("email").send_keys(username)
    driver.find_element_by_id("pass").send_keys(password)
    allure.attach(driver.get_screenshot_as_png(),name="dologin",attachment_type=AttachmentType.png)
